import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
"""
Report: Inherited class for AP/AR reports
functions
    - constructor:
        - arguments: path of the report file
        - use: creates list of strings named data. each list item is a line from the report
"""


def toNum(dataItem):
    numString = dataItem.replace(',', '')
    try:
        num = float(numString)
        return num
    # if ValueError thrown, then item isn't a number and will be returned as is
    except ValueError:
        return dataItem

class DataItem(object):
    def __init__(self):
        self.crTerms = None
        self.accountNum = None
        self.accountName = None
        self.total = None
        self.over90 = None
        self.over60 = None
        self.over30 = None
        self.under30 = None
        self.dueDate = None
        self.effectiveDate = None
        self.invoiceDate = None
        self.invoiceNum = None
        self.voucherNum = None
        self.base = None
        self.coords = []
        self.totalFormula = None

    def getItemInfo(self):
        if not self.base:
            return [self.accountNum, self.accountName, self.crTerms, self.voucherNum, self.invoiceNum, self.invoiceDate,
                    self.effectiveDate, self.dueDate,
                    self.totalFormula, self.under30, self.over30, self.over60, self.over90, self.total]
        else:
            return ['', 'Total:', '', '', '', '', self.totalFormula, self.under30,
                    self.over30, self.over60, self.over90, self.total]

    def setCoords(self, row):
        # columns where all amounts are
        columns = ['I', 'J', 'K', 'L', 'M']
        self.coords = [i+str(row) for i in columns]
        self.totalFormula = '=SUM(' + ','.join(self.coords[1:]) + ')'


class Account(object):
    def __init__(self, openingLine):
        self.accountName = None
        self.accountNum = None
        self.coords = None
        self.setAccountHeader(openingLine)
        self.dataItems = []
        self.coords = None
        self.dataCoords = [[],[],[],[],[]]
        self.formulas = []

    def isComplete(self):
        try:
            return self.dataItems[-1].base
        except IndexError:
            return False

    def getAccountHeader(self):
        return [self.accountNum, self.accountName]

    def setAccountHeader(self, openingLine):
        pass

    def setCoords(self, row):
        # add all data coordinates to formula
        for dataItem in self.dataItems[:-1]:
            for i in range(0, 5):
                self.dataCoords[i].append(dataItem.coords[i])

        # format data coords into formula
        for coord in self.dataCoords:
           self.formulas.append('=SUM('+','.join(coord)+')')

        # set location for formulas
        columns = ['I', 'J', 'K', 'L', 'M']
        self.coords = [i + str(row) for i in columns]

    def getAccountTotal(self):
        return ['', 'Total','', '', '', '','', ''] + self.formulas

    def getAccountSummary(self):
        # returns header + total formulas + hardcoded total (last value of last data item)
        return [''] + self.getAccountHeader() + ['']*3 + self.formulas + [self.dataItems[-1].total]


class Report(object):
    def __init__(self, filePath):
        self.totalAging = None
        self.filePath = filePath
        with open(self.filePath) as file:
            self.data = []
            for line in file:
                self.data.append(line.rstrip('\n'))
            while '' in self.data:
                self.data.remove('')
        file.close()

        self.date = None
        self.accounts = []
        self.coords = []
        self.accountCoords = [[], [], [], [], []]
        self.formulas = []
        self.getDate()

    def checkType(self):
        return None
    def getDate(self):
        dateLine = self.data[0].split(' ')
        while '' in dateLine:
            dateLine.remove('')
        dateIndex = dateLine.index("Date:") + 1  # date will be stored in the index after the "Date:" item
        self.date = dateLine[dateIndex]

    def setCoords(self, row):
        # add all data coordinates to formula
        for account in self.accounts:
            for i in range(0, 5):
                self.accountCoords[i].append(account.coords[i])

        # format data coords into formula
        for coord in self.accountCoords:
            self.formulas.append('=SUM(' + ','.join(coord) + ')')

        # set location for formulas
        columns = ['I', 'J', 'K', 'L', 'M']
        self.coords = [i + str(row) for i in columns]


    def createSpreadsheet(self, name):
        wb = Workbook()
        ws = wb.active

        # header formatting
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15

        for let in 'DEFGHIJKLMO':
            ws.column_dimensions[let].width = 13

        ws.freeze_panes = ws['A5']

        # Header info
        ws['A1'] = ''
        if self.checkType() is 'AP':
            ws['B1'] = 'AP Aging as of Effective Date'
        else:
            ws['B1'] = 'AR Aging as of Effective Date'
        ws['E1'] = self.date

        global currentLine
        currentLine = 4

        def addData(data):
            try:
                data.setCoords(currentLine)
                addRow(data)
            except AttributeError:
                addRow()

        def addRow(data=[''], isDataItem=False):

            global currentLine
            rowCoords = []
            if isinstance(data, DataItem):
                dataList = data.getItemInfo()
                for j in range(0, len(dataList)):
                    cell = ws.cell(row=currentLine, column=j + 1, value=dataList[j])
                    if 5 <= j <= 7:
                        cell.number_format = 'MM/DD/YY'
                    elif j > 7:
                        cell.number_format = '#,##0.00'

            elif isinstance(data, Account):
                dataList = data.getAccountTotal()
                # add hardcoded total to account info (last item in list of base data item)
                hardcodedTotal = data.dataItems[-1].total
                dataList.append(hardcodedTotal)
                for j in range(0, len(dataList)):
                    cell = ws.cell(row=currentLine, column=j + 1, value=dataList[j])
                    if 5 <= j <= 7:
                        cell.number_format = 'MM/DD/YY'
                    elif j > 7:
                        cell.number_format = '#,##0.00'

            else:
                maxCol = len(data)
                for j in range(0, maxCol):
                    ws.cell(row=currentLine, column=j + 1, value=data[j]).number_format = '#,##0.00'
            currentLine += 1
            return rowCoords

        def underlineRow(currentLine, start=1, end=14, grandTotal=False):
            for i in range(start, end):
                if grandTotal:
                    ws.cell(row=currentLine, column=i + 1).border = Border(top=Side(style='double'))
                else:
                    ws.cell(row=currentLine, column=i + 1).border = Border(bottom=Side(style='thin'))


        # header titles
        reportHeader = ['Account #',
                       'Account Name',
                        'cr Terms',
                       'Voucher #',
                       'Invoice #',
                       'Invoice Date',
                       'Effective Date',
                       'Due Date',
                       'Total to Pay',
                       'Under 30',
                       'Over 30',
                       'Over 60',
                       'Over 90',
                       'Total per Account']
        addRow(reportHeader)

        # total coords used to make grand total formula
        for account in self.accounts:
            addRow()
            # iterate through each list item (except for total at end)
            for item in account.dataItems[:-1]:
                addData(item)

            addData(account)
            # close formula string for each element in list
            underlineRow(currentLine - 2)

        # add grand total formula to end of spreadsheet
        self.setCoords(currentLine)

        for i in range(0,3):
            addRow()

        addRow(['', 'Grand Total:', '','','','','','']+self.formulas+[toNum(self.totalAging)])
        underlineRow(currentLine - 1, grandTotal=True)

        # Summary By Account
        for i in range(0,3):
            addRow()

        addRow(['', 'Summary by Account'])
        addRow(['', 'Num', 'Name', '','','', 'Total to pay', 'Under 30', 'Over 30', 'Over 60', 'Over 90', 'Total Per Account'])
        underlineRow(currentLine-1)
        for account in self.accounts:
            addRow(account.getAccountSummary())

        wb.save(name + '.xlsx')
        print(name+' added')