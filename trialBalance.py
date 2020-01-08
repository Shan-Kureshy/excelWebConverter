from openpyxl import Workbook
import re


def getData(rawData):
    # create new list to store all parsed data to
    newData = []

    for line in rawData:
        try:
            # line contains data if line begins at index 3 with a num
            if line[0] == ' ' and re.match('[0-9]', line[1]):

                # split line into list of all its contents
                newLine = line.split('  ')

                # remove leftover garbage list items
                while '' in newLine:
                    newLine.remove('')
                while ' ' in newLine:
                    newLine.remove(' ')
                if '\n' in newLine:
                    newLine.remove('\n')

                newData.append(formatData(newLine))
        except IndexError:
            # if IndexError is thrown, then all data has been collected for that line
            continue

    return newData

# TODO: formatData still producing errors in some lines
def formatData(dataLine):
    formattedData = []

    for item in dataLine:
        index = dataLine.index(item)


        item.rstrip()
        if index > 1:
            item = toNum(item)

        formattedData.append(item)


    # if line has less items, fill with empty items to keep length the same across all lines
    while len(formattedData) < 5:
        formattedData.insert(3, '')

    return formattedData


def toNum(dataItem):
    # remove comma formatting
    numString = dataItem.replace(',', '')

    #remove new line
    if dataItem.endswith('\n'):
        numString = numString[:-1]

    # strings ending with 'cr' need to be turned negative
    negative = False
    if numString.endswith('r'):
        numString = numString[:-2]
        negative = True

    try:
        num = float(numString)
        if negative:
            num *= -1

        return num

    # if ValueError thrown, then item isn't a number and will be returned as is
    except ValueError:
        return dataItem

def createspreadsheet(filePath, spreadsheetName):
    wb = Workbook()
    ws = wb.active

    f = open(filePath)
    contents = f.readlines()

    startDate = ''
    endDate = ''

    # pull date
    count = 0
    while not re.match('[0-9]', contents[8][count]):
        count += 1
    while re.match('[0-9]', contents[8][count]) or contents[8][count] == '/':
        startDate += contents[8][count]
        count += 1
    while not re.match('[0-9]', contents[8][count]):
        count += 1
    while re.match('[0-9]', contents[8][count]) or contents[8][count] == '/':
        endDate += contents[8][count]
        count += 1

    parsedContents = getData(contents)

    # format columns
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 15

    # format spreadsheet intro
    ws['A1'] = '25.15.4'
    ws['C1'] = 'Autogenomics'
    ws['D1'] = 'Trial Balance Summary'
    ws['A3'] = 'Account: '
    ws['B3'] = 'Description: '
    ws['C2'] = 'Beginning Balance:'
    ws['D2'] = 'Period Activity: '
    ws['E2'] = 'Ending Balance:'
    ws['C3'] = startDate
    ws['E3'] = endDate

    dataLength = len(parsedContents)

    # place all data into spreadsheet
    for row in ws.iter_rows(min_row=4, max_col=5, max_row=dataLength + 3):
        for cell in row:
            cell.value = parsedContents[cell.row - 4][cell.column - 1]
            if cell.column >= 3:
                cell.number_format = '#,##0.00'

    # place formula for sum of all data
    beginningVal = '=sum(C4:C' + str(dataLength + 3) + ')'
    periodVal = '=sum(D4:D' + str(dataLength + 3) + ')'
    endingVal = '=sum(E4:E' + str(dataLength + 3) + ')'

    lastRow = ws.cell(row=dataLength + 5, column=3, value='Total:')
    BeginningTotal = ws.cell(row=dataLength + 5, column=3, value=beginningVal)
    BeginningTotal.number_format = '#,##0.00'

    periodTotal = ws.cell(row=dataLength + 5, column=4, value=periodVal)
    periodTotal.number_format = '#,##0.00'

    endingTotal = ws.cell(row=dataLength + 5, column=5, value=endingVal)
    endingTotal.number_format = '#,##0.00'

    # make new sheet for next loop
    ws = wb.create_sheet()
    print(filePath + ' added')

    wb.save(spreadsheetName+'.xlsx')


def main():
    # open workbook

    wb = Workbook()
    ws = wb.active

    # add files to workbook
    while True:

        filePath = input('Enter QAD report file name, or type \'done\': ')
        if filePath == 'done' or filePath == 'Done':
            #remove current sheet since it will be empty
            wb.remove_sheet(ws)
            break
        ws.title = filePath

        # check if file is valid
        try:
            f = open(filePath)
            contents = f.readlines()
            error = False
        except:
            print('Error: Invalid File Name.')
            continue

        startDate = ''
        endDate = ''

        # pull date
        count = 0
        while not re.match('[0-9]', contents[8][count]):
            count += 1
        while re.match('[0-9]', contents[8][count]) or contents[8][count] == '/':
            startDate += contents[8][count]
            count += 1
        while not re.match('[0-9]', contents[8][count]):
            count += 1
        while re.match('[0-9]', contents[8][count]) or contents[8][count] == '/':
            endDate += contents[8][count]
            count += 1

        parsedContents = getData(contents)

        # format columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 15

        # format spreadsheet intro
        ws['A1'] = '25.15.4'
        ws['C1'] = 'Autogenomics'
        ws['D1'] = 'Trial Balance Summary'
        ws['A3'] = 'Account: '
        ws['B3'] = 'Description: '
        ws['C2'] = 'Beginning Balance:'
        ws['D2'] = 'Period Activity: '
        ws['E2'] = 'Ending Balance:'
        ws['C3'] = startDate
        ws['E3'] = endDate

        dataLength = len(parsedContents)

        # place all data into spreadsheet
        for row in ws.iter_rows(min_row=4, max_col=5, max_row=dataLength + 3):
            for cell in row:
                cell.value = parsedContents[cell.row - 4][cell.column - 1]
                if cell.column >= 3:
                    cell.number_format = '#,##0.00'

        # place formula for sum of all data
        beginningVal = '=sum(C4:C' + str(dataLength + 3) + ')'
        periodVal = '=sum(D4:D' + str(dataLength + 3) + ')'
        endingVal = '=sum(E4:E' + str(dataLength + 3) + ')'

        lastRow = ws.cell(row=dataLength + 5, column=3, value='Total:')
        BeginningTotal = ws.cell(row=dataLength + 5, column=3, value=beginningVal)
        BeginningTotal.number_format = '#,##0.00'

        periodTotal = ws.cell(row=dataLength + 5, column=4, value=periodVal)
        periodTotal.number_format = '#,##0.00'

        endingTotal = ws.cell(row=dataLength + 5, column=5, value=endingVal)
        endingTotal.number_format = '#,##0.00'

        #make new sheet for next loop
        ws = wb.create_sheet()
        print(filePath + ' added')

    # close file and save spreadsheet
    try:
        f.close()
        fileName = input('Enter a name to save spreadsheet as: ')
        wb.save(fileName + '.xlsx')
        print('Spreadsheet saved\n')
    except:
        print('No files converted')


if __name__ == "__main__":
    main()
