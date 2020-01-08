import sys
from datetime import datetime
import re
from openpyxl import Workbook
import os.path


def getData(rawData, isCustom):
    # create new list to store all parsed data to
    newData = []
    # TODO: refactor regex into modular chunks
    # TODO: test regex on more test cases, add whitespace in groups

    '''
	dateGroup = r'(?P<date>[0-9]{2}\/[0-9]{2}\/[0-9]{2})'
	acctNumGroup = r'(?P<acctNum>\w+\s)'
	addressGroup = r'(?P<address>NO ADDR|[^ ]+)'
	nameGroup = r'(?P<name>(?:[^ ]+\s)+\s)'
	num2Group = r'(?P<num2>\w{2,}\s)?'
	prefixGroup = r'(?P<prefix>\s[a-zA-Z]\s)?'
	num3Group = r'(?P<num3>\w{2,}\s)?'
	debitCreditGroup = r'(?P<debit>[0-9,]+.[0-9]{2}\s)\s*(?P<credit>[0-9,]+.[0-9]{2})'
	space = r'\s*'
	customRegex = dateGroup+space+acctNumGroup+space+addressGroup+space+\
				  nameGroup+space+num2Group+space+prefixGroup+space+num3Group+space+debitCreditGroup
	defaultRegex = dateGroup+space+acctNumGroup+space+addressGroup+space+\
				  num2Group+space+prefixGroup+space+num3Group+space+debitCreditGroup
	journalCustomRegex = dateGroup+space+acctNumGroup+space+addressGroup+space+\
				  nameGroup+space+debitCreditGroup
	journalDefaultRegex = dateGroup+space+acctNumGroup+space+\
				  nameGroup+space+debitCreditGroup
	
	customRegex = r'\s*(?P<date>[0-9]{2}\/[0-9]{2}\/[0-9]{2})\s*(?P<acctNum>\w+\s)\s*(?P<address>NO ADDR|[^ ]+\s)\s*(?P<name>(?:[^ ]+\s)+)\s*(?P<num2>\s\w{2,}\s)?\s*(?P<prefix>\s[a-zA-Z]\s)?\s*(?P<num3>\w{2,}\s)?\s*(?P<debit>[0-9,]+.[0-9]{2}\s)\s*(?P<credit>[0-9,]+.[0-9]{2})'
	defaultRegex = r'\s*(?P<date>[0-9]{2}\/[0-9]{2}\/[0-9]{2})\s*(?P<acctNum>\w+\s)\s+(?P<address>NO ADDR|[^ ]+\s)\s*(?P<num2>\w{2,}\s>)?\s*(?P<prefix>\s[a-zA-Z]\s)?\s*(?P<num3>\w{2,}\s)?\s*(?P<debit>[0-9,]+.[0-9]{2}\s)\s*(?P<credit>[0-9,]+.[0-9]{2})'
	journalCustomRegex = r'\s*(?P<date>[0-9]{2}\/[0-9]{2}\/[0-9]{2})\s*(?P<acctNum>\w+\s)\s+(?P<address>NO ADDR|[^ ]+\s)\s*(?P<name>(?:[^ ]+\s{1,2})+)\s*(?P<debit>[0-9,]+.[0-9]{2}\s)\s*(?P<credit>[0-9,]+.[0-9]{2})'
	journalDefaultRegex = r'\s*(?P<date>[0-9]{2}\/[0-9]{2}\/[0-9]{2})\s*(?P<acctNum>\w+\s)\s*(?P<name>(?:[^ ]+\s{1,2})+)\s*(?P<debit>[0-9,]+.[0-9]{2}\s)\s*(?P<credit>[0-9,]+.[0-9]{2})'
	'''

    for line in rawData:
        newLine = line
        try:
            isJournal = re.match(r'^.* (?:JL|RV)[0-9]{5,}.*$', newLine)
            # line contains data if line begins at index 3 with a num
            if newLine[2] == ' ' and re.match('[0-9]', newLine[3]):

                # add extra space after name to prevent errors
                spaceMatch = re.search(r' [0-9]{5} ', newLine)
                if spaceMatch:
                    pos = spaceMatch.regs[0][0]
                    newLine = newLine[:pos] + ' ' + newLine[pos:]

                # determine which regex expression to use
                if isCustom:
                    if isJournal:
                        match = re.match(
                            r'\s*(?P<date>[0-9]{2}\/[0-9]{2}\/[0-9]{2})\s*(?P<acctNum>[^ ]+\s)\s+(?P<address>NO ADDR|[^ ]+\s)\s*(?P<name>(?:[^ ]+\s{1,2})+)\s*(?P<debit>[0-9,]+.[0-9]{2}\s)\s*(?P<credit>[0-9,]+.[0-9]{2})',
                            newLine)
                        date = datetime.strptime(match.group('date'), '%m/%d/%y')
                        newLine = [date,
                                   match.group('acctNum'),
                                   match.group('name'),
                                   '', '', '',
                                   toNum(match.group('debit')), toNum(match.group('credit'))]

                    else:
                        match = re.match(
                            r'\s*(?P<date>[0-9]{2}\/[0-9]{2}\/[0-9]{2})\s*(?P<acctNum>[^ ]+\s)\s*(?P<address>NO ADDR|[^ ]+\s)\s*(?P<name>(?:[^ ]+\s{1,2})+)\s*(?P<num2>[^ ]{2,}\s)?\s*(?P<prefix>\s[a-zA-Z]\s)?\s*(?P<num3>[^ ]{2,}\s)?\s*(?P<debit>[0-9,]+.[0-9]{2}\s)\s*(?P<credit>[0-9,]+.[0-9]{2})',
                            newLine)

                        date = datetime.strptime(match.group('date'), '%m/%d/%y')
                        newLine = [date, match.group('acctNum'),
                                   match.group('name'),
                                   match.group('num2'),
                                   match.group('prefix') if match.group('prefix') else '',
                                   match.group('num3'),
                                   toNum(match.group('debit')),
                                   toNum(match.group('credit'))]
                else:
                    if isJournal:
                        match = re.match(
                            r'\s*(?P<date>[0-9]{2}\/[0-9]{2}\/[0-9]{2})\s*(?P<acctNum>[^ ]+\s)\s*(?P<name>(?:[^ ]+\s{1,2})+)\s*(?P<debit>[0-9,]+.[0-9]{2}\s)\s*(?P<credit>[0-9,]+.[0-9]{2})',
                            newLine)
                        date = datetime.strptime(match.group('date'), '%m/%d/%y')
                        newLine = [date,
                                   match.group('acctNum'),
                                   match.group('name'),
                                   '', '', '',
                                   toNum(match.group('debit')),
                                   toNum(match.group('credit'))]
                    else:
                        match = re.match(
                            r'\s*(?P<date>[0-9]{2}\/[0-9]{2}\/[0-9]{2})\s*(?P<acctNum>[^ ]+\s)\s+(?P<address>NO ADDR|[^ ]+\s)\s*(?P<num2>[^ ]{2,}\s)?\s*(?P<prefix>\s[a-zA-Z]\s)?\s*(?P<num3>[^ ]{2,}\s)?\s*(?P<debit>[0-9,]+.[0-9]{2}\s)\s*(?P<credit>[0-9,]+.[0-9]{2})',
                            newLine)
                        date = datetime.strptime(match.group('date'), '%m/%d/%y')
                        newLine = [date,
                                   match.group('acctNum'),
                                   match.group('num2'),
                                   match.group('prefix') if match.group('prefix') else '',
                                   match.group('num3'), '',
                                   toNum(match.group('debit')),
                                   toNum(match.group('credit'))]

                # newLine = errorCheck(newLine, newLine)
                newData.append(newLine)


        except IndexError as e:
            # if IndexError is thrown, then all data has been collected for that line
            if e.args[0] == 'no such group':
                errorLine = lineError(newLine)
                newData.append(errorLine)
            continue
        except AttributeError:
            errorLine = lineError(newLine)
            newData.append(errorLine)

    return newData


def errorCheck(lineList, lineString):
    '''
	errorDict = {
		1: r'^[a-zA-Z]{2}[0-9]{5,15}$',
		3: r'^[0-9]*$',
		4: r'\w*',
		5: r'\w*',
		6: r'[0-9,]+.[0-9]{0,2}',
		7: r'[0-9,]+.[0-9]{0,2}'
	}

	for key in errorDict:
		if key == 0:
			if not isinstance(lineList[key], datetime):
				return lineError(lineString)
		else:
			if not re.match(errorDict[key], lineList[key]):
				return lineError(lineString)
	return lineList
	'''
    if isinstance(lineList[0], datetime):
        if re.match(r'([a-zA-Z]{2}[0-9]{5,15}\s)?', lineList[1]):
            if re.match(r'([0-9]*\s)?', lineList[3]):
                if re.match(r'[a-zA-Z]?', lineList[4]):
                    if re.match(r'[0-9,]+.[0-9]{0,2}', '{0:.2f}'.format(lineList[6])):
                        if re.match(r'[0-9,]+.[0-9]{2}', '{0:.2f}'.format(lineList[7])):
                            return lineList
    return lineError(lineString)


def lineError(line):
    errorLine = [''] * 8
    errorLine[0] = '(Error converting line. Please enter manually ) ' + line
    return errorLine


def getDate(rawData):
    for line in rawData:
        if re.match(r'^.*(?<!Total )Activity To Date:.*$', line):
            dates = re.findall(r'[0-9]{2}/[0-9]{2}/[0-9]{2}', line)
            break

    return dates


def getHeader(rawData):
    headerLine = rawData[9]

    header = headerLine.split('  ')
    while '' in header:
        header.remove('')

    # last list item is irrelevant, only return first 2
    return header[:2]


def getBalances(rawData):
    balances = []
    numBalances = []

    for line in rawData:
        if line[:5] == 'Total':
            balanceVal = line.split(':')
            balanceVal[1] = balanceVal[1].strip()
            balances.append(balanceVal[1])

    # turn strings into numbers

    for item in balances:
        numBalances.append(toNum(item))

    return numBalances


def toNum(dataItem):
    numString = dataItem.replace(',', '')
    negative = False
    if numString.endswith('r'):
        numString = numString[:-2]
        negative = True

    try:
        num = float(numString)
        if negative:
            num *= -1

        return num

    # if ValueError thrown, then item isn't a number and will be returned as is
    except ValueError:
        return dataItem


def createSpreadsheet(filePath, spreadsheetName):
    f = open(filePath)

    contents = f.readlines()
    isCustom = re.search(r'^.*25\.16\.2.*$', contents[0])

    parsedContents = getData(contents, isCustom)

    header = getHeader(contents)

    balances = getBalances(contents)

    dates = getDate(contents)

    # spreadsheetName = input('enter a spreadsheet name: ')
    wb = Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25 if isCustom else 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20

    ws['A1'] = '25.16.2' if isCustom else '25.15.2'
    ws['B1'] = 'Acct. Balance Detail'

    # display dates
    ws['C2'] = 'Start:'
    ws['D2'] = 'End:'
    ws['C3'] = dates[0]
    ws['C3'].number_format = 'MM/DD/YY'
    ws['D3'] = dates[1]
    ws['D3'].number_format = 'MM/DD/YY'

    # display acct. info
    ws['A3'] = header[0]
    ws['B3'] = header[1]

    def addRow(currentLine, data=None, maxCol=8):
        if data is None:
            data = [''] * 8
            
        ws.cell(row=currentLine, column=1).number_format = 'MM/DD/YY'
        ws.cell(row=currentLine, column=7).number_format = '#,##0.00'
        ws.cell(row=currentLine, column=8).number_format = '#,##0.00'
        for i in range(0, maxCol):
            ws.cell(row=currentLine, column=i + 1, value=data[i])

    currentLine = 7
    for line in parsedContents:
        addRow(currentLine, line)
        currentLine += 1

    # display closing info
    ws.cell(row=currentLine, column=7, value='Beginning Balance:')
    ws.cell(row=currentLine, column=8, value=balances[0]).number_format = '#,##0.00'
    currentLine += 1

    ws.cell(row=currentLine, column=7, value='Period Total:')
    ws.cell(row=currentLine, column=8, value=balances[1]).number_format = '#,##0.00'
    currentLine += 1

    ws.cell(row=currentLine, column=7, value='Ending Balance:')
    ws.cell(row=currentLine, column=8, value=balances[2]).number_format = '#,##0.00'
    currentLine += 1

    f.close()
    wb.save(spreadsheetName + '.xlsx')


def main():
    filePath = input('Enter a QAD file name: ')

    while not os.path.isfile(filePath):
        if filePath == 'skureshy':
            __runTest__()
            sys.exit()

        if os.path.isfile(filePath + '.txt'):
            filePath += '.txt'
        else:
            filePath = input(
                'Error: File not found. Please make sure file is in the correct folder, and name is being entered correctly.'
                '\nEnter a QAD file name: ')

    spreadsheetName = input('Enter a name to save spreadsheet as: ')
    createSpreadsheet(filePath, spreadsheetName)


def __runTest__():
    filePaths = ['test/3', 'test/4', 'test/ad1', 'test/ad2', 'test/broken', 'test/broken2',
                 'test/cash.txt', 'test/custom1.txt', 'test/custom2.txt', 'test/custom3.txt']
    for path in filePaths:
        createSpreadsheet(path, path)


if __name__ == '__main__':
    main()
